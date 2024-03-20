import openpyxl
from openpyxl.utils import get_column_letter
from faker import Faker
from random import choice

fake = Faker()

def save_to_excel(data, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['First Name', 'Last Name', 'Gender', 'Nationality', 'Language', 'TRN', 'Address Name', 'Type', 'Emirate', 'Description Location', 'Location', 'Date Inquiry', 'Services', 'Source', 'Description', 'Status'])
    
    for row in data:
        ws.append(row)

    for col in range(1, len(ws[1]) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 20

    wb.save(filename)

def run_and_save_to_excel():
    # Generate fake data
    data = []
    for _ in range(1000):
        first_name = fake.first_name()
        last_name = fake.last_name()
        gender = choice(['male', 'female'])
        nationality = choice(["Unknown", "arabic expat male", "arabic expat female", "company", "UAE male", "UAE female",
                    "Euroupe&american&Australia&Russia male", "Euroupe&american&Australia&Russia female",
                    "asia&africa male", "asia&africa Female"])
        language = choice(["Arabic", "French", "English", "Spanish"])
        trn = fake.uuid4()
        address_name = fake.street_name()
        address_type = choice(['house', 'company'])
        emirate = choice(["Emara 1", "Emara 2", "Emara 3"])
        description_location = fake.text()
        location = fake.address()
        date_inq = fake.date_this_year()
        service = choice(["electric", "climatisation", "plumber"])
        source = choice(["from service provider", "website", "Direct booking", "from points", "affiliate marketing",
                "inbound calls", "outbound calls", "Instagram", "facebook", "linkedin", "tiktok", "snapchat",
                "telegram", "recommended customer", "whats up broadcast", "whats up marketing", "sms campaign",
                "email marketing"])
        description = fake.text()
        status1 = choice(["new", "connecting", "pending", "cancel", "underproccess", "send Q or B"])
        status = choice([status1,'done'])

        data.append([
            first_name, last_name, gender, nationality, language, trn,
            address_name, address_type, emirate, description_location, location,
            date_inq, service, source, description, status
        ])

    # Save data to Excel file
    save_to_excel(data, 'generated_data.xlsx')

# Run the script and save data to Excel
run_and_save_to_excel()
